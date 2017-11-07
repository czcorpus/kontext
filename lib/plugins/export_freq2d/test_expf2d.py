# Copyright (c) 2017 Charles University, Faculty of Arts,
#                    Institute of the Czech National Corpus
# Copyright (c) 2017 Tomas Machalek <tomas.machalek@gmail.com>
# Copyright (c) 2017 Petr Duda <petrduda@seznam.cz>
#
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; version 2
# dated June, 1991.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

"""
Unittests for the export_freq2d plugin
"""
import unittest
import sys
import os

from openpyxl import load_workbook

from plugins.export_freq2d.default_xlsx import XLSXExport

sys.path.insert(0, os.path.realpath('%s/../../../scripts/' % os.path.dirname(os.path.realpath(__file__))))

import autoconf
import initializer
import plugins


class ExpFreq2dTest(unittest.TestCase):
    def __init__(self, *args, **kwargs):
        super(ExpFreq2dTest, self).__init__(*args, **kwargs)

        tmp = '/tmp/xls_test/'
        if not os.path.exists(tmp):
            os.makedirs(tmp)

        self.sample = {"attr1": "lemma_lc", "attr2": "div.group",
                       "labels1": ["friday", "monday", "saturday", "sunday", "thursday", "tuesday", "wednesday"],
                       "labels2": ["Acquis", "Core", "Europarl", "PressEurop", "Subtitles", "Syndicate"],
                       "alphaLevel": "0.05",
                       "minAbsFreq": 1, "data": [
                [[5.55, 6.15, 6.87], [24.06, 26.7, 29.66], [11.98, 13.29, 14.81], [40.4, 44.83, 50.11],
                 [28.88, 32.05, 35.59],
                 [2.43, 3.42, 5.47]],
                [[8.27, 9.18, 10.22], [20.16, 22.37, 24.86], [17.94, 19.91, 22.16], [41.56, 46.12, 51.54],
                 [23.83, 26.44, 29.36], [0.79, 1.33, 2.81]],
                [[3.6, 3.99, 4.47], [33.75, 37.45, 41.6], [5.6, 6.21, 6.96], [22.38, 24.83, 27.92],
                 [28.27, 31.37, 34.82],
                 [1.2, 1.9, 3.63]],
                [[4.7, 5.22, 5.83], [49.6, 55.04, 61.12], [9.13, 10.13, 11.3], [60.74, 67.41, 75.16],
                 [31.78, 35.26, 39.15],
                 [5.36, 6.64, 8.66]],
                [[3.99, 4.43, 4.95], [14.74, 16.35, 18.19], [35.15, 39.01, 43.35], [27.32, 30.32, 34],
                 [16.94, 18.8, 20.88],
                 [0.39, 0.76, 2.41]],
                [[5.06, 5.61, 6.27], [13.78, 15.29, 17], [13.84, 15.36, 17.11], [24.12, 26.77, 30.06],
                 [16.08, 17.85, 19.82],
                 [0.65, 1.14, 2.75]],
                [[5.03, 5.58, 6.23], [12.85, 14.26, 15.86], [21.98, 24.4, 27.14], [21.22, 23.54, 26.49],
                 [12.79, 14.19, 15.76],
                 [0.52, 0.95, 2.62]]]}

        settings = autoconf.settings
        logger = autoconf.logger

        initializer.init_plugin('export_freq2d', module=plugins.export_freq2d)
        loader = plugins.runtime.EXPORT_FREQ2D.instance
        self.expf2d = loader.load_plugin("xlsx")

    def test_content_type(self):
        self.assertTrue(self.expf2d.content_type(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_set_content(self):
        """
        save content using the old_save() method
        """
        self.expf2d.set_content(self.sample.get("attr1"), self.sample.get("attr2"), self.sample.get("labels1"),
                                self.sample.get("labels2"), self.sample.get("alphaLevel"),
                                self.sample.get("minAbsFreq"),
                                self.sample.get("data"))
        file_path = self.expf2d.old_save()
        # check whether file created
        self.assertTrue(os.path.isfile(file_path))

        # ----------
        # check data
        # ----------
        wb = load_workbook(file_path)
        sheet = wb.active
        labels1 = self.sample.get("labels1")
        labels2 = self.sample.get("labels2")

        # find first row of data table by looking for the first row label in the first column
        for r in range(1, sheet.max_row):
            if sheet.cell(row=r, column=1).internal_value == labels1[0]:
                start_row = r
                break

        # check whether labels1 match
        row_ind = 0
        for label in labels1:
            self.assertEqual(sheet.cell(row=start_row+row_ind, column=1).internal_value, label)
            row_ind += 1

        # check whether labels2 match
        col_ind = 2
        for label in labels2:
            self.assertEqual(sheet.cell(row=start_row - 1, column=col_ind).internal_value, label)
            col_ind += 3

        # check whether data cells match
        sample_data = self.sample.get("data")
        row_ind = 0
        for doc_row in sample_data:
            col_ind = 2
            for triad in doc_row:
                for val in triad:
                    self.assertEqual(val, sheet.cell(row=start_row+row_ind, column=col_ind).internal_value)
                    col_ind += 1
            row_ind += 1        


if __name__ == '__main__':
    unittest.main()
