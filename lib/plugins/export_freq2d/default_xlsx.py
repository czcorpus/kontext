# Copyright (c) 2017 Charles University, Faculty of Arts,
#                    Institute of the Czech National Corpus
# Copyright (c) 2017 Tomas Machalek <tomas.machalek@gmail.com>
# Copyright (c) 2017 Petr Duda <petrduda@seznam.cz>
#
# This program is free software; you can redistribute it and/or
# modify it under the terms of the GNU General Public License
# as published by the Free Software Foundation; version 2
# dated June, 1991.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.

# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston, MA
# 02110-1301, USA.

from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Side, Color, PatternFill, Font, Border, colors
from plugins.export_freq2d import AbstractExportFreq2d

# define styles
KONTEXT_GREEN = "d2edc0"
GREEN_FILL = PatternFill(patternType='solid', fgColor=Color(KONTEXT_GREEN))
GRAY_FONT = Font(color="999999")
GREEN_BORDER = Side(style='medium', color=KONTEXT_GREEN)
NO_BORDER = Side(style='thin', color=colors.WHITE)


class XLSXExport(AbstractExportFreq2d):

    def __init__(self):
        self._wb = Workbook()

    def content_type(self):
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def set_content(self, attr1, attr2, labels1, labels2, alpha_level, min_freq, min_freq_type, data):
        tbl = XLSExportInternal2d(self._wb)
        tbl.set_content(attr1=attr1, attr2=attr2, labels1=labels1, labels2=labels2, alpha_level=alpha_level,
                        min_freq=min_freq, min_freq_type=min_freq_type, data=data)

    def set_content_flat(self, headings, alpha_level, min_freq, min_freq_type, data):
        tbl = XLSExportInternal2d_flat(self._wb)
        tbl.set_content(headings=headings, alpha_level=alpha_level, min_freq=min_freq, min_freq_type=min_freq_type,
                        data=data)

    def raw_content(self):
        output = StringIO()
        self._wb.save(filename=output)
        return output.getvalue()


class XLSExportInternal2d_flat(object):
    def __init__(self, wb):
        self._wb = wb
        self._sheet = self._wb.active

    def set_content(self, headings, alpha_level, min_freq, min_freq_type, data):
        # --------------
        # render caption
        # --------------
        table_width = len(headings[0]) + 1
        self._sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table_width)
        self._sheet["A1"].fill = GREEN_FILL
        self._sheet["A1"] = "Two-attribute interrelationship"
        self._sheet["A1"].font = Font(bold=True)

        self._sheet["A3"] = "Alpha level:"
        self._sheet["B3"] = float(alpha_level)
        self._sheet["A4"] = 'Min. ({0}):'.format(min_freq_type)
        self._sheet["B4"] = float(min_freq)

        # render headings - hard-coded to allow for specific merged cells layout
        headings_row = 6
        column_index = 2

        for col_val in headings:
            act_col = self._sheet.cell(row=headings_row, column=column_index)
            act_col.value = col_val
            if column_index == 4:
                column_index += 3  # will be merged
            else:
                column_index += 1

        self._sheet.merge_cells(start_row=headings_row, start_column=4,
                                end_row=headings_row, end_column=6)
        self._sheet.merge_cells(start_row=headings_row, start_column=7,
                                end_row=headings_row, end_column=9)
        self._sheet.cell(row=headings_row, column=4).alignment = Alignment(horizontal='center')
        self._sheet.cell(row=headings_row, column=7).alignment = Alignment(horizontal='center')

        for col in range(1, table_width):
            self._sheet.cell(row=headings_row, column=col).fill = GREEN_FILL
            self._sheet.cell(row=headings_row, column=col).font = Font(bold=True)

        # -----------------
        # render data table
        # -----------------
        data_first_row = headings_row
        # fill in data
        row_index = data_first_row + 1
        grey_cols = [4, 6, 7, 9]
        max_width_attr_1 = len(headings[0]) + 2
        max_width_attr_2 = len(headings[1]) + 2
        for row_val in data:
            # print row no.
            self._sheet.cell(row=row_index, column=1).value = str(row_index - data_first_row) + "."
            column_index = 2
            for col_val in row_val:
                act_col = self._sheet.cell(row=row_index, column=column_index)
                if col_val is not None:
                    act_col.value = col_val
                column_index += 1
            # range columns in gray
            for col in grey_cols:
                self._sheet.cell(row=row_index, column=col).font = GRAY_FONT
            row_index += 1

        # get max len of attr cols
        for col in data:
            if len(col[0]) > max_width_attr_1:
                max_width_attr_1 = len(col[0])
            if len(col[1]) > max_width_attr_2:
                max_width_attr_2 = len(col[1])

        # set width of attr cols
        self._sheet.column_dimensions["B"].width = max_width_attr_1
        self._sheet.column_dimensions["C"].width = max_width_attr_2


class XLSExportInternal2d(object):
    def __init__(self, wb):
        self._wb = wb
        self._sheet = self._wb.active

    def set_content(self, attr1, attr2, labels1, labels2, alpha_level, min_freq, min_freq_type, data):
        # --------------
        # render caption
        # --------------
        table_width = len(data[0]) * 3 + 1
        self._sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table_width)
        self._sheet["A1"].fill = GREEN_FILL
        self._sheet["A1"] = "Two-attribute interrelationship: " + attr1 + "\\" + attr2
        self._sheet["A1"].font = Font(bold=True)
        self._sheet["A3"] = "Attr. 1:"
        self._sheet["B3"] = attr1
        self._sheet["A4"] = "Attr. 2:"
        self._sheet["B4"] = attr2
        self._sheet["A5"] = "Alpha level:"
        self._sheet["B5"] = float(alpha_level)
        self._sheet["A6"] = 'Min. freq. ({0}):'.format(min_freq_type)
        self._sheet["B6"] = min_freq

        # -----------------
        # render data table
        # -----------------
        data_first_row = 8

        # fill attr cell
        self._sheet.cell(row=data_first_row, column=1).value = attr1 + "\\" + attr2
        self._sheet.cell(row=data_first_row, column=1).font = Font(bold=True)

        # fill in row names
        rows = labels1
        row_index = data_first_row + 1
        for row_name in rows:
            r = self._sheet.cell(row=row_index, column=1)
            r.value = row_name
            r.fill = GREEN_FILL
            r.alignment = Alignment(horizontal='right')
            row_index += 1

        # fill in column names
        columns = labels2
        column_index = 2
        for col_name in columns:
            c = self._sheet.cell(row=data_first_row, column=column_index)
            c.value = col_name
            c.fill = GREEN_FILL
            c.alignment = Alignment(horizontal='center')
            self._sheet.merge_cells(start_row=data_first_row, start_column=column_index, end_row=data_first_row,
                                    end_column=column_index + 2)
            column_index += 3

        # fill in data
        row_index = data_first_row + 1
        for row_val in data:
            column_index = 2
            for col_val in row_val:
                min_col = self._sheet.cell(row=row_index, column=column_index)
                act_col = self._sheet.cell(row=row_index, column=column_index + 1)
                max_col = self._sheet.cell(row=row_index, column=column_index + 2)
                min_col.border = Border(bottom=GREEN_BORDER, right=NO_BORDER)
                act_col.border = Border(bottom=GREEN_BORDER, right=NO_BORDER)
                max_col.border = Border(right=GREEN_BORDER, bottom=GREEN_BORDER)
                if col_val is not None:
                    min_col.value = col_val[0]
                    act_col.value = col_val[1]
                    max_col.value = col_val[2]
                    min_col.font = GRAY_FONT
                    max_col.font = GRAY_FONT
                column_index += 3
            row_index += 1


def create_instance():
    return XLSXExport()
